#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
import json
import os

def excel_to_html(excel_file, output_file='output.html'):
    """将 Excel 文件转换为 HTML，保留公式"""
    
    # 先读取公式
    wb_formulas = openpyxl.load_workbook(excel_file, data_only=False)
    # 再读取计算结果
    wb_values = openpyxl.load_workbook(excel_file, data_only=True)
    
    # 存储所有工作表的数据
    sheets_data = {}
    
    for sheet_name in wb_formulas.sheetnames:
        sheet_formulas = wb_formulas[sheet_name]
        sheet_values = wb_values[sheet_name]
        
        sheet_data = {
            'name': sheet_name,
            'rows': [],
            'formulas': {},
            'max_row': sheet_formulas.max_row,
            'max_col': sheet_formulas.max_column
        }
        
        # 遍历所有单元格
        for row_idx, row_formulas in enumerate(sheet_formulas.iter_rows(min_row=1, max_row=sheet_formulas.max_row, 
                                                       min_col=1, max_col=sheet_formulas.max_column), 1):
            row_values = list(sheet_values.iter_rows(min_row=row_idx, max_row=row_idx, 
                                                     min_col=1, max_col=sheet_formulas.max_column))[0]
            row_data = []
            
            for col_idx, (cell_formula, cell_value) in enumerate(zip(row_formulas, row_values), 1):
                cell_key = f"{row_idx}_{col_idx}"
                
                # 检查是否有公式
                if cell_formula.data_type == 'f':  # formula
                    formula = cell_formula.value
                    # 获取公式的计算值
                    if cell_value.value is None:
                        calculated_value = ""
                    else:
                        calculated_value = str(cell_value.value)
                    
                    sheet_data['formulas'][cell_key] = {
                        'formula': formula,
                        'row': row_idx,
                        'col': col_idx
                    }
                    row_data.append({
                        'value': calculated_value,
                        'formula': formula,
                        'has_formula': True
                    })
                else:
                    # 普通单元格
                    if cell_formula.value is None:
                        cell_value_str = ""
                    else:
                        cell_value_str = str(cell_formula.value)
                    
                    row_data.append({
                        'value': cell_value_str,
                        'formula': None,
                        'has_formula': False
                    })
            
            sheet_data['rows'].append(row_data)
        
        sheets_data[sheet_name] = sheet_data
    
    # 生成 HTML
    html_content = generate_html(sheets_data)
    
    # 写入文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"HTML 文件已生成: {output_file}")
    return output_file

def generate_html(sheets_data):
    """生成 HTML 内容"""
    
    # 将数据转换为 JSON 以便 JavaScript 使用
    sheets_json = json.dumps(sheets_data, ensure_ascii=False, indent=2)
    
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel 公式转换网页</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            padding: 20px;
        }}
        
        h1 {{
            color: #333;
            margin-bottom: 20px;
            border-bottom: 2px solid #4CAF50;
            padding-bottom: 10px;
        }}
        
        .tabs {{
            display: flex;
            border-bottom: 2px solid #e0e0e0;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }}
        
        .tab {{
            padding: 12px 24px;
            cursor: pointer;
            background: #f9f9f9;
            border: none;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
            margin-right: 5px;
            font-size: 14px;
            transition: all 0.3s;
        }}
        
        .tab:hover {{
            background: #e8f5e9;
        }}
        
        .tab.active {{
            background: #4CAF50;
            color: white;
        }}
        
        .sheet-container {{
            display: none;
            overflow-x: auto;
        }}
        
        .sheet-container.active {{
            display: block;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
            font-size: 13px;
        }}
        
        th, td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
            min-width: 80px;
        }}
        
        th {{
            background-color: #4CAF50;
            color: white;
            font-weight: bold;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        
        td {{
            background-color: white;
        }}
        
        td.formula-cell {{
            background-color: #fff9c4;
            font-style: italic;
        }}
        
        td.formula-cell:hover {{
            background-color: #fff59d;
        }}
        
        td.editable {{
            cursor: text;
        }}
        
        td.editing {{
            background-color: #e3f2fd !important;
            outline: 2px solid #2196F3;
        }}
        
        td:focus {{
            outline: 2px solid #2196F3;
        }}
        
        .formula-info {{
            position: absolute;
            background: #333;
            color: white;
            padding: 5px 10px;
            border-radius: 4px;
            font-size: 12px;
            display: none;
            z-index: 1000;
            max-width: 300px;
            word-wrap: break-word;
        }}
        
        .col-header {{
            background-color: #4CAF50;
            color: white;
            font-weight: bold;
            text-align: center;
        }}
        
        .row-header {{
            background-color: #4CAF50;
            color: white;
            font-weight: bold;
            text-align: center;
            min-width: 50px;
        }}
        
        .info-panel {{
            background: #e3f2fd;
            padding: 15px;
            border-radius: 4px;
            margin-bottom: 20px;
            border-left: 4px solid #2196F3;
        }}
        
        .info-panel h3 {{
            margin-bottom: 10px;
            color: #1976D2;
        }}
        
        .formula-display {{
            font-family: 'Courier New', monospace;
            background: #f5f5f5;
            padding: 2px 6px;
            border-radius: 3px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Excel 公式转换网页</h1>
        
        <div class="info-panel">
            <h3>使用说明</h3>
            <p>• 点击标签页切换不同的工作表</p>
            <p>• 黄色背景的单元格包含公式，鼠标悬停可查看公式</p>
            <p>• 所有单元格都可以编辑，修改数据后公式会自动重新计算</p>
            <p>• 双击单元格进入编辑模式，按 Enter 或点击其他地方完成编辑</p>
        </div>
        
        <div class="tabs" id="tabs"></div>
        
        <div id="sheets-container"></div>
    </div>
    
    <script>
        const sheetsData = {sheets_json};
        const cellValues = {{}}; // 存储所有单元格的值
        const cellFormulas = {{}}; // 存储所有单元格的公式
        const formulaDependencies = {{}}; // 存储公式依赖关系
        
        // 初始化页面
        function init() {{
            const tabsContainer = document.getElementById('tabs');
            const sheetsContainer = document.getElementById('sheets-container');
            
            // 创建标签页和工作表
            Object.keys(sheetsData).forEach((sheetName, index) => {{
                // 创建标签
                const tab = document.createElement('button');
                tab.className = 'tab' + (index === 0 ? ' active' : '');
                tab.textContent = sheetName;
                tab.onclick = () => switchTab(sheetName);
                tabsContainer.appendChild(tab);
                
                // 创建工作表容器
                const sheetDiv = document.createElement('div');
                sheetDiv.className = 'sheet-container' + (index === 0 ? ' active' : '');
                sheetDiv.id = `sheet-${{sheetName}}`;
                
                const table = createTable(sheetsData[sheetName], sheetName);
                sheetDiv.appendChild(table);
                sheetsContainer.appendChild(sheetDiv);
            }});
            
            // 初始化所有公式
            recalculateAllFormulas();
        }}
        
        function switchTab(sheetName) {{
            // 切换标签
            document.querySelectorAll('.tab').forEach(tab => {{
                tab.classList.remove('active');
            }});
            event.target.classList.add('active');
            
            // 切换工作表
            document.querySelectorAll('.sheet-container').forEach(container => {{
                container.classList.remove('active');
            }});
            document.getElementById(`sheet-${{sheetName}}`).classList.add('active');
        }}
        
        function createTable(sheetData, sheetName) {{
            const table = document.createElement('table');
            
            // 创建表头行（列号）
            const headerRow = document.createElement('tr');
            const emptyCell = document.createElement('th');
            emptyCell.className = 'row-header';
            headerRow.appendChild(emptyCell);
            
            for (let col = 1; col <= sheetData.max_col; col++) {{
                const th = document.createElement('th');
                th.className = 'col-header';
                th.textContent = getColumnLetter(col);
                headerRow.appendChild(th);
            }}
            table.appendChild(headerRow);
            
            // 创建数据行
            sheetData.rows.forEach((row, rowIndex) => {{
                const tr = document.createElement('tr');
                
                // 行号
                const rowHeader = document.createElement('th');
                rowHeader.className = 'row-header';
                rowHeader.textContent = rowIndex + 1;
                tr.appendChild(rowHeader);
                
                // 数据单元格
                row.forEach((cell, colIndex) => {{
                    const td = document.createElement('td');
                    const cellId = `${{sheetName}}_${{rowIndex + 1}}_${{colIndex + 1}}`;
                    const cellRef = getColumnLetter(colIndex + 1) + (rowIndex + 1);
                    
                    // 存储初始值
                    cellValues[cellId] = cell.value || '';
                    
                    // 设置单元格属性
                    td.setAttribute('data-cell-id', cellId);
                    td.setAttribute('data-cell-ref', cellRef);
                    td.setAttribute('data-sheet', sheetName);
                    td.setAttribute('data-row', rowIndex + 1);
                    td.setAttribute('data-col', colIndex + 1);
                    td.className = 'editable';
                    td.contentEditable = true;
                    
                    if (cell.has_formula) {{
                        td.className += ' formula-cell';
                        cellFormulas[cellId] = cell.formula;
                        // 格式化公式结果
                        const formattedValue = formatNumber(cell.value);
                        td.textContent = formattedValue;
                        cellValues[cellId] = formattedValue;
                        td.title = `公式: ${{cell.formula}}`;
                        // 公式单元格不可直接编辑（但可以通过修改引用的单元格来改变）
                        td.contentEditable = false;
                        td.style.cursor = 'default';
                    }} else {{
                        // 格式化普通单元格的数值
                        const formattedValue = formatNumber(cell.value);
                        td.textContent = formattedValue;
                        cellValues[cellId] = formattedValue;
                    }}
                    
                    // 添加编辑事件
                    td.addEventListener('focus', function() {{
                        this.classList.add('editing');
                        if (!this.hasAttribute('data-original-value')) {{
                            this.setAttribute('data-original-value', this.textContent);
                        }}
                    }});
                    
                    td.addEventListener('blur', function() {{
                        this.classList.remove('editing');
                        let newValue = this.textContent.trim();
                        const cellId = this.getAttribute('data-cell-id');
                        const oldValue = cellValues[cellId] || '';
                        
                        // 格式化输入的数字
                        const formattedValue = formatNumber(newValue);
                        this.textContent = formattedValue;
                        newValue = formattedValue;
                        
                        if (newValue !== oldValue) {{
                            cellValues[cellId] = newValue;
                            // 如果这个单元格有依赖它的公式，重新计算
                            if (formulaDependencies[cellId]) {{
                                // 递归更新所有依赖的公式
                                const updatedCells = new Set();
                                function updateDependencies(cellId) {{
                                    if (updatedCells.has(cellId)) return;
                                    updatedCells.add(cellId);
                                    
                                    if (formulaDependencies[cellId]) {{
                                        formulaDependencies[cellId].forEach(formulaCellId => {{
                                            recalculateFormula(formulaCellId);
                                            updateDependencies(formulaCellId);
                                        }});
                                    }}
                                }}
                                updateDependencies(cellId);
                            }}
                        }}
                        this.removeAttribute('data-original-value');
                    }});
                    
                    td.addEventListener('keydown', function(e) {{
                        if (e.key === 'Enter') {{
                            e.preventDefault();
                            this.blur();
                        }} else if (e.key === 'Escape') {{
                            const originalValue = this.getAttribute('data-original-value');
                            this.textContent = originalValue || '';
                            this.blur();
                        }}
                    }});
                    
                    td.onmouseenter = function(e) {{
                        if (cell.has_formula) {{
                            showFormulaTooltip(e, cell.formula);
                        }}
                    }};
                    
                    td.onmouseleave = function() {{
                        if (cell.has_formula) {{
                            hideFormulaTooltip();
                        }}
                    }};
                    
                    tr.appendChild(td);
                }});
                
                table.appendChild(tr);
            }});
            
            return table;
        }}
        
        function getColumnLetter(colNum) {{
            let result = '';
            while (colNum > 0) {{
                colNum--;
                result = String.fromCharCode(65 + (colNum % 26)) + result;
                colNum = Math.floor(colNum / 26);
            }}
            return result;
        }}
        
        let tooltip = null;
        
        function showFormulaTooltip(event, formula) {{
            if (!tooltip) {{
                tooltip = document.createElement('div');
                tooltip.className = 'formula-info';
                document.body.appendChild(tooltip);
            }}
            
            tooltip.textContent = `公式: ${{formula}}`;
            tooltip.style.display = 'block';
            tooltip.style.left = (event.pageX + 10) + 'px';
            tooltip.style.top = (event.pageY + 10) + 'px';
        }}
        
        function hideFormulaTooltip() {{
            if (tooltip) {{
                tooltip.style.display = 'none';
            }}
        }}
        
        // 格式化数字，保留2位小数
        function formatNumber(value) {{
            if (value === '' || value === null || value === undefined) return '';
            
            // 尝试转换为数字
            const numValue = parseFloat(value);
            if (isNaN(numValue)) {{
                // 如果不是数字，返回原值
                return value.toString();
            }}
            
            // 格式化数字，保留2位小数
            return numValue.toFixed(2);
        }}
        
        // 解析公式中的单元格引用
        function parseFormulaDependencies(formula) {{
            if (!formula || !formula.startsWith('=')) return [];
            
            const dependencies = [];
            // 匹配 Excel 单元格引用，如 A1, B10, C2 等
            const cellRefRegex = /([A-Z]+)([0-9]+)/g;
            let match;
            
            while ((match = cellRefRegex.exec(formula)) !== null) {{
                const col = match[1];
                const row = parseInt(match[2]);
                // 需要找到对应的 cellId，这里先返回引用字符串
                dependencies.push(match[0]);
            }}
            
            return dependencies;
        }}
        
        // 获取单元格的值
        function getCellValue(sheetName, cellRef) {{
            // 解析单元格引用，如 A1 -> row=1, col=1
            const colMatch = cellRef.match(/([A-Z]+)/);
            const rowMatch = cellRef.match(/([0-9]+)/);
            
            if (!colMatch || !rowMatch) return 0;
            
            const col = colMatch[1];
            const row = parseInt(rowMatch[1]);
            const colNum = columnLetterToNumber(col);
            
            const cellId = `${{sheetName}}_${{row}}_${{colNum}}`;
            
            // 先尝试从 DOM 获取当前值
            const cell = document.querySelector(`[data-cell-id="${{cellId}}"]`);
            if (cell) {{
                const value = cell.textContent.trim();
                const numValue = parseFloat(value);
                return isNaN(numValue) ? 0 : numValue;
            }}
            
            // 如果 DOM 中没有，从 cellValues 获取
            const value = cellValues[cellId] || '';
            const numValue = parseFloat(value);
            return isNaN(numValue) ? 0 : numValue;
        }}
        
        // 将列字母转换为数字 (A=1, B=2, ..., Z=26, AA=27, ...)
        function columnLetterToNumber(letters) {{
            let result = 0;
            for (let i = 0; i < letters.length; i++) {{
                result = result * 26 + (letters.charCodeAt(i) - 64);
            }}
            return result;
        }}
        
        // 重新计算公式单元格
        function recalculateFormula(cellId) {{
            const formula = cellFormulas[cellId];
            if (!formula) return;
            
            const cell = document.querySelector(`[data-cell-id="${{cellId}}"]`);
            if (!cell) return;
            
            const sheetName = cell.getAttribute('data-sheet');
            
            try {{
                // 将 Excel 公式转换为 JavaScript 表达式
                let jsFormula = formula.substring(1); // 去掉开头的 =
                
                // 替换单元格引用为实际值
                const cellRefRegex = /([A-Z]+)([0-9]+)/g;
                jsFormula = jsFormula.replace(cellRefRegex, (match, col, row) => {{
                    const cellRef = col + row;
                    return getCellValue(sheetName, cellRef);
                }});
                
                // 计算表达式
                const result = Function('"use strict"; return (' + jsFormula + ')')();
                
                // 格式化结果并更新单元格显示
                const formattedResult = formatNumber(result);
                cell.textContent = formattedResult;
                cellValues[cellId] = formattedResult;
            }} catch (e) {{
                console.error('公式计算错误:', formula, e);
                cell.textContent = '#ERROR';
            }}
        }}
        
        // 重新计算所有公式
        function recalculateAllFormulas() {{
            // 清空依赖关系
            Object.keys(formulaDependencies).forEach(key => delete formulaDependencies[key]);
            
            // 建立所有单元格引用到 cellId 的映射
            const refToCellId = {{}};
            document.querySelectorAll('[data-cell-ref]').forEach(cell => {{
                const sheetName = cell.getAttribute('data-sheet');
                const cellRef = cell.getAttribute('data-cell-ref');
                const cellId = cell.getAttribute('data-cell-id');
                refToCellId[`${{sheetName}}_${{cellRef}}`] = cellId;
            }});
            
            // 建立公式依赖关系
            Object.keys(cellFormulas).forEach(cellId => {{
                const formula = cellFormulas[cellId];
                const cell = document.querySelector(`[data-cell-id="${{cellId}}"]`);
                if (!cell) return;
                
                const sheetName = cell.getAttribute('data-sheet');
                const deps = parseFormulaDependencies(formula);
                
                deps.forEach(dep => {{
                    const depCellId = refToCellId[`${{sheetName}}_${{dep}}`];
                    if (depCellId) {{
                        if (!formulaDependencies[depCellId]) {{
                            formulaDependencies[depCellId] = [];
                        }}
                        if (!formulaDependencies[depCellId].includes(cellId)) {{
                            formulaDependencies[depCellId].push(cellId);
                        }}
                    }}
                }});
            }});
            
            // 计算所有公式
            Object.keys(cellFormulas).forEach(cellId => {{
                recalculateFormula(cellId);
            }});
        }}
        
        // 页面加载时初始化
        window.onload = init;
    </script>
</body>
</html>"""
    
    return html

if __name__ == '__main__':
    excel_file = '1.xlsx'
    if os.path.exists(excel_file):
        excel_to_html(excel_file, 'index.html')
    else:
        print(f"错误: 找不到文件 {excel_file}")
